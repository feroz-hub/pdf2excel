#!/usr/bin/env python3
"""Verify NIST SP 800-53 Standard Assessment workbook output."""

from __future__ import annotations

import re
import sys
from collections import Counter

from openpyxl import load_workbook

_FRONT_MATTER_TITLES = frozenset({"authority", "abstract", "errata"})
bad_patterns_text = [
    "This publication is available free of charge",
    "doi.org",
    "CHAPTER THREE PAGE",
    "APPENDIX C PAGE",
]
valid_id_re = re.compile(r"^[A-Z]{2}-\d+(\(\d+\))?$")


def verify(path: str) -> None:
    wb = load_workbook(path, data_only=False)
    ws = wb["Standard Assessment"]
    rows = []
    for row_idx in range(10, ws.max_row + 1):
        clause_id = str(ws.cell(row_idx, 2).value or "").strip()
        title = str(ws.cell(row_idx, 3).value or "").strip()
        text = str(ws.cell(row_idx, 4).value or "").strip()
        classification = str(ws.cell(row_idx, 5).value or "").strip()
        if clause_id or title or text:
            rows.append((row_idx, clause_id, title, text, classification))

    blank_clause = [r for r in rows if not r[1]]
    invalid_ids = [r for r in rows if r[1] and not valid_id_re.match(r[1])]
    bad_rows = [
        r for r in rows
        if "joint task force" in (r[1] + " " + r[2]).lower()
        or "nist sp 800-53, r ev" in (r[1] + " " + r[2]).lower()
        or r[2].strip().lower() in _FRONT_MATTER_TITLES
        or any(p.lower() in r[3].lower() for p in bad_patterns_text)
    ]
    one_letter = [r for r in rows if len(r[3]) <= 1]
    duplicates = [k for k, v in Counter(r[1] for r in rows if r[1]).items() if v > 1]

    print("Total rows:", len(rows))
    print("First row:", rows[0][:4] if rows else None)
    print("Blank clause IDs:", len(blank_clause))
    print("Invalid IDs:", len(invalid_ids))
    print("Bad pattern rows:", len(bad_rows))
    print("One-letter rows:", len(one_letter))
    print("Duplicate clause IDs:", len(duplicates))
    print("First 10 IDs:", [r[1] for r in rows[:10]])

    assert rows, "No data rows found"
    assert rows[0][1] == "AC-1", f"First row should be AC-1, got {rows[0][1]}"
    assert len(rows) < 3000, f"Too many rows, likely old extractor used: {len(rows)}"
    assert len(blank_clause) == 0, "Blank clause IDs found"
    assert len(bad_rows) == 0, "Bad front matter/header/footer rows found"
    assert len(one_letter) == 0, "One-letter rows found"
    assert len(duplicates) == 0, "Duplicate clause IDs found"
    print("NIST output verification passed.")


def main() -> int:
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <workbook.xlsx>", file=sys.stderr)
        return 2
    verify(sys.argv[1])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
