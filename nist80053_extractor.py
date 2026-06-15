"""NIST SP 800-53 Rev. 5 extractor — one Standard Assessment row per control / enhancement.

This module parses the NIST SP 800-53 Rev. 5 PDF (a structured control catalog)
into clean Standard Assessment items.  The correct extraction unit is:

    one NIST base control  →  one row
    one control enhancement  →  one row

Non-control content (cover, TOC, chapters 1–2, errata, references, appendices,
rotated side-margin text, page headers/footers, DOI boilerplate) is excluded.

Public API
----------
    is_nist_800_53(pdf_path)               → bool
    extract_nist_800_53_items(pdf_path)     → list[dict]
    write_nist_review_workbook(items, path) → None
"""

from __future__ import annotations

import logging
import re
from typing import Any, Dict, List, Optional

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

log = logging.getLogger("pdf2excel.nist80053")

# ---------------------------------------------------------------------------
# Constants / patterns
# ---------------------------------------------------------------------------

# Regex: base control heading — e.g. "AC-1 POLICY AND PROCEDURES"
BASE_CONTROL_RE = re.compile(
    r"^(?P<id>[A-Z]{2}-\d+)\s+(?P<title>[A-Z0-9][A-Z0-9 ,/\-\u2014:&()]+)$"
)

# Regex: control enhancement — e.g. "(1) AUTOMATED SYSTEM ACCOUNT MANAGEMENT"
# The title may contain a pipe separator for the enhancement subtitle.
ENHANCEMENT_RE = re.compile(
    r"^\((?P<num>\d+)\)\s+(?P<title>[A-Z0-9][A-Z0-9 ,/\-\u2014:&()|]+)$"
)

# Family section heading — e.g. "3.1 ACCESS CONTROL"
FAMILY_HEADING_RE = re.compile(
    r"^3\.\d+\s+[A-Z][A-Z ]+$"
)

# Withdrawn marker
WITHDRAWN_RE = re.compile(r"\[Withdrawn\s*:", re.IGNORECASE)

# Boilerplate / noise patterns
_DOI_FRAGMENT = "doi.org/10.6028/NIST.SP.800-53r5"
_DOI_TEXT = "This publication is available free of charge"
_PAGE_HEADER = "NIST SP 800-53, REV. 5"
_UNDERLINE = "_" * 30

# Page-level markers for catalog boundaries
_CHAPTER_THREE_MARKERS = ("CHAPTER THREE PAGE", )
_END_MARKERS = ("REFERENCES PAGE", "APPENDIX A PAGE", "APPENDIX B PAGE", "APPENDIX C PAGE")

# Requirement language for classification
_REQUIREMENT_RE = re.compile(
    r"\b(shall not|shall|must|is required to|are required to)\b", re.IGNORECASE
)

# Soft-hyphen and bad-encoding cleanup
_SOFT_HYPHEN = "\u00ad"
_CLEANUP_RE = re.compile(r"[\u00ad\ufeff]")

# Side-margin x-threshold (rotated DOI text sits at x0 ≈ 18)
_SIDE_MARGIN_X = 45.0
# Header y-threshold (page header text at top ≈ 38-48)
_HEADER_Y = 55.0
# Footer y-threshold (page footer at bottom ≈ 740+)
_FOOTER_Y = 740.0


# ---------------------------------------------------------------------------
# Detection
# ---------------------------------------------------------------------------

def is_nist_800_53(pdf_path: str) -> bool:
    """Return True if the PDF appears to be NIST SP 800-53 Rev. 5."""
    if pdfplumber is None:
        return False
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) < 50:
                return False
            # Check first 5 pages for NIST SP 800-53 identification
            for page in pdf.pages[:5]:
                text = page.extract_text() or ""
                if "800-53" in text and "Revision 5" in text:
                    return True
                if "800-53" in text and "Rev. 5" in text:
                    return True
                if "NIST SP 800-53" in text:
                    return True
            return False
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Text cleaning helpers
# ---------------------------------------------------------------------------

def _clean_text(text: str) -> str:
    """Normalize whitespace, remove soft hyphens and BOM characters."""
    text = _CLEANUP_RE.sub("", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _is_boilerplate(line: str) -> bool:
    """Return True if the line is page header/footer/DOI boilerplate."""
    stripped = line.strip()
    if not stripped:
        return True
    if _DOI_FRAGMENT in stripped:
        return True
    if _DOI_TEXT in stripped:
        return True
    if stripped.startswith(_PAGE_HEADER):
        return True
    if stripped.startswith(_UNDERLINE):
        return True
    # Page number lines: "CHAPTER THREE PAGE NN" or just a number
    if re.match(r"^(CHAPTER THREE|CHAPTER TWO|CHAPTER ONE|REFERENCES|APPENDIX\s+[A-C])\s+PAGE\s+\d+$", stripped, re.IGNORECASE):
        return True
    if re.match(r"^\d{1,3}$", stripped):
        return True
    if re.match(r"^[ivxlc]+$", stripped, re.IGNORECASE):
        return True
    # Single broken characters from rotated text
    if len(stripped) <= 2 and not re.match(r"^[A-Z]{2}$", stripped):
        return True
    return False


def _is_side_margin_word(word: dict) -> bool:
    """Return True if a word is rotated side-margin text or DOI fragment."""
    if word.get("upright") is False:
        return True
    # Words hugging the left margin that are part of the rotated DOI text
    if word["x0"] < _SIDE_MARGIN_X:
        text = word["text"]
        # Single chars or short fragments at the margin are DOI debris
        if len(text) <= 3:
            return True
        if any(frag in text.lower() for frag in [
            "publication", "available", "charge", "https", "doi.org"
        ]):
            return True
    return False


# ---------------------------------------------------------------------------
# Page text extraction (word-level, filtering noise)
# ---------------------------------------------------------------------------

def _extract_clean_lines(page: Any) -> List[dict]:
    """Extract cleaned lines from a page, filtering rotated/boilerplate text.

    Returns a list of dicts: {text, top, size, font, is_bold}.
    Each dict represents one visual line (words grouped by vertical position).
    """
    try:
        words = page.extract_words(
            use_text_flow=False,
            keep_blank_chars=False,
            extra_attrs=["size", "fontname", "upright"],
        )
    except Exception:
        words = page.extract_words(extra_attrs=["size", "fontname"])

    # Filter out rotated / side-margin words
    clean_words = []
    for w in words:
        if _is_side_margin_word(w):
            continue
        # Skip header region
        if w["top"] < _HEADER_Y:
            continue
        # Skip footer region
        if w["top"] > _FOOTER_Y:
            continue
        clean_words.append(w)

    if not clean_words:
        return []

    # Group words into lines by vertical position (tolerance ~3pt)
    clean_words.sort(key=lambda w: (round(w["top"] / 3) * 3, w["x0"]))
    lines = []
    current_top = None
    current_words = []

    for w in clean_words:
        top_bucket = round(w["top"] / 3) * 3
        if current_top is None or top_bucket != current_top:
            if current_words:
                lines.append(_make_line(current_words))
            current_top = top_bucket
            current_words = [w]
        else:
            current_words.append(w)
    if current_words:
        lines.append(_make_line(current_words))

    # Filter boilerplate lines
    result = []
    for line in lines:
        if _is_boilerplate(line["text"]):
            continue
        result.append(line)

    return result


def _make_line(words: List[dict]) -> dict:
    """Merge a cluster of words into a single line dict."""
    words.sort(key=lambda w: w["x0"])
    text = " ".join(w["text"] for w in words)
    text = _clean_text(text)
    # Use the most common font size / name in the line
    sizes = [w.get("size", 10) for w in words]
    fonts = [w.get("fontname", "") for w in words]
    avg_size = sum(sizes) / len(sizes) if sizes else 10.0
    # Pick the dominant font
    font_counts: Dict[str, int] = {}
    for f in fonts:
        font_counts[f] = font_counts.get(f, 0) + 1
    dominant_font = max(font_counts, key=font_counts.get) if font_counts else ""
    is_bold = "Bold" in dominant_font or "bold" in dominant_font

    return {
        "text": text,
        "top": words[0]["top"],
        "size": avg_size,
        "font": dominant_font,
        "is_bold": is_bold,
        "x0": words[0]["x0"],
    }


# ---------------------------------------------------------------------------
# Control heading detection
# ---------------------------------------------------------------------------

def _is_base_control_heading(line: dict) -> Optional[re.Match]:
    """Check if the line is a base control heading like 'AC-1 POLICY AND PROCEDURES'."""
    if not line["is_bold"] or line["size"] < 10.5:
        return None
    m = BASE_CONTROL_RE.match(line["text"])
    return m


def _is_enhancement_heading(line: dict) -> Optional[re.Match]:
    """Check if the line is an enhancement heading like '(1) TITLE | SUBTITLE'."""
    m = ENHANCEMENT_RE.match(line["text"])
    return m


def _is_family_heading(line: dict) -> bool:
    """Check if the line is a family section heading like '3.1 ACCESS CONTROL'."""
    return line["size"] >= 13.0 and bool(FAMILY_HEADING_RE.match(line["text"]))


# ---------------------------------------------------------------------------
# Catalog boundary detection
# ---------------------------------------------------------------------------

def _find_catalog_pages(pdf) -> tuple:
    """Return (start_page_idx, end_page_idx) of the Chapter Three control catalog.

    start is the first page whose header says 'CHAPTER THREE PAGE'
    end is the first page whose header says 'REFERENCES PAGE' (exclusive).
    """
    start_idx = None
    end_idx = len(pdf.pages)

    for pg_idx, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        lines = text.strip().split("\n")[:5]
        for line in lines:
            stripped = line.strip().upper()
            if start_idx is None:
                for marker in _CHAPTER_THREE_MARKERS:
                    if marker in stripped:
                        start_idx = pg_idx
                        break
            if start_idx is not None:
                for marker in _END_MARKERS:
                    if marker in stripped:
                        end_idx = pg_idx
                        return (start_idx, end_idx)

    return (start_idx or 0, end_idx)


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------

def extract_nist_800_53_items(pdf_path: str) -> List[dict]:
    """Extract NIST SP 800-53 Rev. 5 controls and enhancements.

    Returns Standard Assessment items:
    {
        "clause_id": "AC-1",
        "title": "POLICY AND PROCEDURES",
        "text": "Control: ... Discussion: ... Related Controls: ...",
        "classification": "Requirement" | "Information",
        "source_page": 45,
        "confidence": 0.95,
        "issues": []
    }
    """
    if pdfplumber is None:
        raise RuntimeError("pdfplumber is required. Run: pip install pdfplumber")

    items: List[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        start_idx, end_idx = _find_catalog_pages(pdf)
        log.info(
            "Catalog pages: %d to %d (of %d total)",
            start_idx + 1, end_idx, len(pdf.pages),
        )

        # First pass: collect all lines across the catalog pages
        all_lines: List[dict] = []  # each has extra "page" key
        for pg_idx in range(start_idx, end_idx):
            page = pdf.pages[pg_idx]
            page_num = pg_idx + 1
            lines = _extract_clean_lines(page)
            for line in lines:
                line["page"] = page_num
            all_lines.extend(lines)

    # Second pass: parse into controls and enhancements
    items = _parse_controls(all_lines)

    # Validation pass
    items = _validate_items(items)

    return items


def _parse_controls(all_lines: List[dict]) -> List[dict]:
    """Parse cleaned lines into control/enhancement items."""
    items: List[dict] = []

    current_base_id: Optional[str] = None
    current_base_title: Optional[str] = None
    current_family: Optional[str] = None

    # Accumulator for the current control/enhancement
    current_id: Optional[str] = None
    current_title: Optional[str] = None
    current_text_lines: List[str] = []
    current_page: Optional[int] = None
    current_is_enhancement = False

    def _flush():
        """Emit the accumulated control/enhancement as an item."""
        nonlocal current_id, current_title, current_text_lines, current_page
        if current_id is None:
            return
        text = _clean_text(" ".join(current_text_lines))
        if not text:
            current_id = None
            current_title = None
            current_text_lines = []
            current_page = None
            return

        is_withdrawn = bool(WITHDRAWN_RE.search(text))
        classification = "Information" if is_withdrawn else "Requirement"

        # For non-withdrawn controls, check for obligation language as extra validation
        if not is_withdrawn and not _REQUIREMENT_RE.search(text):
            # Even if no shall/must, NIST controls are requirements by definition
            classification = "Requirement"

        issues = []
        confidence = 0.95

        # Basic quality checks
        if len(text) < 20:
            issues.append("text_too_short")
            confidence = 0.6
        if _DOI_FRAGMENT in text:
            issues.append("contains_doi_fragment")
            confidence = 0.3
        if is_withdrawn:
            confidence = 0.90

        items.append({
            "clause_id": current_id,
            "title": current_title or "",
            "text": text,
            "classification": classification,
            "source_page": current_page,
            "confidence": confidence,
            "issues": issues,
        })

        current_id = None
        current_title = None
        current_text_lines = []
        current_page = None

    skip_intro = True  # Skip initial Chapter Three prose before first control

    for line in all_lines:
        text = line["text"]

        # Skip family section headings (e.g. "3.1 ACCESS CONTROL")
        if _is_family_heading(line):
            current_family = text
            continue

        # Skip the "Quick link to ..." lines
        if text.startswith("Quick link to"):
            continue

        # Check for base control heading
        m_base = _is_base_control_heading(line)
        if m_base:
            skip_intro = False
            _flush()
            current_base_id = m_base.group("id")
            current_base_title = m_base.group("title").strip()
            current_id = current_base_id
            current_title = current_base_title
            current_text_lines = []
            current_page = line["page"]
            current_is_enhancement = False
            continue

        # Check for enhancement heading
        m_enh = _is_enhancement_heading(line)
        if m_enh and current_base_id:
            _flush()
            enh_num = m_enh.group("num")
            enh_title = m_enh.group("title").strip()
            current_id = f"{current_base_id}({enh_num})"
            # The raw enhancement heading already includes the family name
            # with pipe separator (e.g. "ACCOUNT MANAGEMENT | AUTOMATED ...").
            # Use it directly to avoid duplicating the base title.
            current_title = enh_title
            current_text_lines = []
            current_page = line["page"]
            current_is_enhancement = True
            continue

        if skip_intro:
            continue

        # Accumulate text for the current control/enhancement
        if current_id is not None:
            current_text_lines.append(text)

    # Flush the last accumulated control
    _flush()

    return items


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate_items(items: List[dict]) -> List[dict]:
    """Run quality checks on extracted items and flag issues."""
    seen_ids: Dict[str, int] = {}

    for item in items:
        clause_id = item.get("clause_id", "")
        issues = item.get("issues", [])

        # Duplicate check
        if clause_id in seen_ids:
            issues.append("duplicate_clause_id")
            item["confidence"] = min(item.get("confidence", 1.0), 0.5)
        seen_ids[clause_id] = seen_ids.get(clause_id, 0) + 1

        # Missing clause_id
        if not clause_id:
            issues.append("missing_clause_id")
            item["confidence"] = min(item.get("confidence", 1.0), 0.3)

        # Missing title
        if not item.get("title"):
            issues.append("missing_title")
            item["confidence"] = min(item.get("confidence", 1.0), 0.5)

        # Text too short
        text = item.get("text", "")
        if len(text) < 10:
            issues.append("text_too_short")
            item["confidence"] = min(item.get("confidence", 1.0), 0.3)

        # Contains DOI
        if _DOI_FRAGMENT in text:
            issues.append("contains_doi_url")
            item["confidence"] = min(item.get("confidence", 1.0), 0.2)

        # Single character text
        if len(text.strip()) <= 2:
            issues.append("single_char_text")
            item["confidence"] = min(item.get("confidence", 1.0), 0.1)

        # Page header in text
        if _PAGE_HEADER in text:
            issues.append("contains_page_header")
            item["confidence"] = min(item.get("confidence", 1.0), 0.2)

        # TOC dot leaders
        if re.search(r"\.{4,}", text):
            issues.append("toc_dot_leader")
            item["confidence"] = min(item.get("confidence", 1.0), 0.2)

        item["issues"] = issues

    # Check total count — NIST 800-53 should have hundreds of controls
    if len(items) < 50:
        log.warning("Only %d items extracted — expected hundreds for NIST 800-53", len(items))

    return items


# ---------------------------------------------------------------------------
# Review workbook
# ---------------------------------------------------------------------------

def write_nist_review_workbook(items: List[dict], out_path: str) -> None:
    """Write a review workbook with extraction details."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "NIST 800-53 Review"

    headers = [
        "source_page", "clause_id", "title", "classification",
        "text_preview", "text_length", "confidence", "issues",
    ]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    body_font = Font(name="Arial", size=9)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for item in items:
        text = item.get("text", "")
        preview = text[:200] + "..." if len(text) > 200 else text
        ws.append([
            item.get("source_page", ""),
            item.get("clause_id", ""),
            item.get("title", ""),
            item.get("classification", ""),
            preview,
            len(text),
            item.get("confidence", 1.0),
            ", ".join(item.get("issues", [])),
        ])

    # Column widths
    widths = [10, 12, 30, 14, 60, 10, 10, 25]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Apply body formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(out_path)
    log.info("Review workbook written: %s (%d items)", out_path, len(items))
