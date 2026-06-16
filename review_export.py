"""Review export module — writes rejected and low-confidence rows for analyst review."""

from __future__ import annotations

from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import validation


def write_review_workbook(
    items: List[Dict[str, Any]],
    out_path: str,
    profile_name: str = "unknown",
) -> None:
    """Write review workbook with export status and issue details."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Items"

    headers = [
        "source_page", "source_type", "clause_id", "title", "classification",
        "text_preview", "text_length", "confidence", "issues", "raw_text",
        "bbox", "profile", "export_status",
    ]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    body_font = Font(name="Arial", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for item in items:
        text = item.get("text") or item.get("raw_text") or ""
        issues_list = item.get("issues") or []
        issues_str = ", ".join(issues_list) if isinstance(issues_list, list) else str(issues_list)
        bbox = item.get("bbox") or ""
        has_blocking = any(i in validation.BLOCKING_ISSUES for i in issues_list)
        export_status = item.get("export_status")
        if not export_status:
            export_status = "rejected" if has_blocking else "exported"

        ws.append([
            item.get("page") or item.get("source_page") or "",
            item.get("source_type") or item.get("block_type") or "paragraph",
            item.get("clause_id") or "",
            item.get("title") or "",
            item.get("classification") or "Information",
            text[:150] + ("..." if len(text) > 150 else ""),
            len(text),
            item.get("confidence", 1.0),
            issues_str,
            text[:500],
            str(bbox) if bbox else "",
            profile_name,
            export_status,
        ])

        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = wrap_align if col_idx in (4, 5, 6, 9, 10) else top_align

    widths = {
        1: 12, 2: 12, 3: 12, 4: 20, 5: 15,
        6: 50, 7: 12, 8: 12, 9: 25, 10: 40, 11: 20, 12: 15, 13: 14,
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)
