"""Unit tests for the hybrid extraction pipeline.

Run with:
    pytest tests/test_hybrid.py
"""

from __future__ import annotations

import os
import sys
import tempfile
from unittest.mock import MagicMock, patch

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import preflight
import extract_blocks
import validate_extraction
import router
from standard_export import write_standard_assessment, blocks_to_items


def test_preflight_page_classification():
    # Mock a page with mostly text words
    page = MagicMock()
    page.width = 600
    page.height = 800
    page.images = []
    page.rects = []
    page.lines = []
    page.curves = []
    
    # 1. Prose page
    page.extract_words.return_value = [
        {"x0": 50, "x1": 150, "top": 100, "bottom": 115, "text": "This", "size": 10, "fontname": "Arial"},
        {"x0": 160, "x1": 250, "top": 100, "bottom": 115, "text": "is", "size": 10, "fontname": "Arial"},
        {"x0": 260, "x1": 350, "top": 100, "bottom": 115, "text": "text.", "size": 10, "fontname": "Arial"},
    ]
    page.find_tables.return_value = []
    page.extract_text.return_value = "This is text."
    
    meta = preflight.analyze_page(page, page_num=2)
    assert meta["page_type"] == "prose"
    assert not meta["is_scanned"]
    assert not meta["is_multi_column"]

    # 2. Table page
    table_mock = MagicMock()
    table_mock.bbox = [50, 200, 550, 400]
    page.find_tables.return_value = [table_mock]
    
    meta = preflight.analyze_page(page, page_num=2)
    assert meta["page_type"] == "table"


def test_block_layout_sorting():
    # Test column sorting: left column top-to-bottom, then right column
    blocks = [
        {"page": 1, "block_type": "paragraph", "text": "Right column top", "bbox": [320, 100, 550, 150]},
        {"page": 1, "block_type": "paragraph", "text": "Left column top", "bbox": [50, 100, 280, 150]},
        {"page": 1, "block_type": "paragraph", "text": "Left column bottom", "bbox": [50, 200, 280, 250]},
        {"page": 1, "block_type": "paragraph", "text": "Right column bottom", "bbox": [320, 200, 550, 250]},
    ]
    sorted_blocks = extract_blocks.sort_blocks_layout(blocks, page_width=600, is_multi_column=True)
    
    assert sorted_blocks[0]["text"] == "Left column top"
    assert sorted_blocks[1]["text"] == "Left column bottom"
    assert sorted_blocks[2]["text"] == "Right column top"
    assert sorted_blocks[3]["text"] == "Right column bottom"


def test_validation_checks():
    items = [
        {"clause_id": "1.2", "title": "Requirements", "text": "The system shall log errors.", "classification": "Requirement", "confidence": 1.0},
        # Broken clause order (1.1 follows 1.2)
        {"clause_id": "1.1", "title": "Scope", "text": "Short scope.", "classification": "Requirement", "confidence": 1.0},
        # Naked page number
        {"clause_id": "1.3", "title": "Boilerplate", "text": "Page 15 of 20", "classification": "Information", "confidence": 1.0},
        # TOC dot leader
        {"clause_id": "1.4", "title": "TOC", "text": "Chapter 1.....................15", "classification": "Information", "confidence": 1.0},
        # Duplicate row
        {"clause_id": "1.5", "title": "Repeated", "text": "This is a repeated paragraph to test duplicate checker.", "classification": "Requirement", "confidence": 1.0},
        {"clause_id": "1.6", "title": "Repeated Again", "text": "This is a repeated paragraph to test duplicate checker.", "classification": "Requirement", "confidence": 1.0},
    ]
    
    validated = validate_extraction.validate_items(items)
    
    # Check issue assignments
    assert "broken_clause_order" in validated[1]["issues"]
    assert "page_number_included" in validated[2]["issues"]
    assert "toc_dot_leader" in validated[3]["issues"]
    assert "duplicate_row" in validated[4]["issues"]
    assert "duplicate_row" in validated[5]["issues"]
    
    # Confidence should be lowered on bad matches
    assert validated[2]["confidence"] <= 0.2
    assert validated[3]["confidence"] <= 0.3


def test_standard_export_issues_sheet():
    items = [
        {"clause_id": "1.1", "title": "Obligation", "text": "The developer must document code.", "classification": "Requirement", "confidence": 0.4, "issues": ["text_confidence_low"]},
        {"clause_id": "1.2", "title": "Info", "text": "Page 10", "classification": "Information", "confidence": 0.2, "issues": ["page_number_included"]},
    ]
    
    with tempfile.TemporaryDirectory() as tmpdir:
        out = os.path.join(tmpdir, "test_issues.xlsx")
        write_standard_assessment(items, out)
        
        # Verify Excel sheets
        wb = openpyxl.load_workbook(out)
        assert "Extraction_Issues" in wb.sheetnames
        
        ws = wb["Extraction_Issues"]
        assert ws["A1"].value == "Row"
        assert ws["G1"].value == "Issues"
        
        # Row 2 (first issue)
        assert ws["E2"].value == "The developer must document code."
        assert "text_confidence_low" in ws["G2"].value
        
        # Row 3 (second issue)
        assert ws["E3"].value == "Page 10"
        assert "page_number_included" in ws["G3"].value


def test_router_hybrid_flow():
    # Mock PDF Page & Words
    page = MagicMock()
    page.width = 600
    page.height = 800
    page.images = []
    page.rects = []
    page.lines = []
    page.curves = []
    page.find_tables.return_value = []
    page.extract_words.return_value = [
        {"x0": 50, "x1": 200, "top": 100, "bottom": 115, "text": "Article 1 Obligations", "size": 12, "fontname": "Times-Bold"},
        {"x0": 50, "x1": 400, "top": 130, "bottom": 145, "text": "The operator shall verify safety.", "size": 10, "fontname": "Times"},
    ]
    page.extract_text.return_value = "Article 1 Obligations\nThe operator shall verify safety."
    
    pdf_mock = MagicMock()
    pdf_mock.pages = [page]
    
    with patch("pdfplumber.open") as mock_open:
        mock_open.return_value.__enter__.return_value = pdf_mock
        
        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "router_out.xlsx")
            
            # Call convert in hybrid 'auto' mode
            result = router.convert(
                source="dummy.pdf",
                out_path=out,
                mode="auto",
                fmt="standard",
                standard_id="MLSR",
                standard_title="Test standard",
                skip_cover=False
            )
            
            assert result.mode == "auto"
            assert result.fmt == "standard"
            assert len(result.items) > 0
            assert result.items[0]["clause_id"] == "Article 1"
            assert result.items[0]["title"] == "Obligations"
            assert result.items[0]["text"] == "The operator shall verify safety."
            assert result.items[0]["classification"] == "Requirement"
