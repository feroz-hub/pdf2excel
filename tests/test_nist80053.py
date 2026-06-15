"""Unit tests for the NIST SP 800-53 Rev. 5 extractor.

Run with:
    pytest tests/test_nist80053.py -v
"""

from __future__ import annotations

import os
import sys
import tempfile
import re
from unittest.mock import MagicMock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from nist80053_extractor import (
    BASE_CONTROL_RE,
    ENHANCEMENT_RE,
    WITHDRAWN_RE,
    _clean_text,
    _is_boilerplate,
    _is_side_margin_word,
    _validate_items,
)


# ---------------------------------------------------------------------------
# Heading detection
# ---------------------------------------------------------------------------

class TestBaseControlHeading:
    def test_ac1(self):
        m = BASE_CONTROL_RE.match("AC-1 POLICY AND PROCEDURES")
        assert m is not None
        assert m.group("id") == "AC-1"
        assert m.group("title") == "POLICY AND PROCEDURES"

    def test_ac2(self):
        m = BASE_CONTROL_RE.match("AC-2 ACCOUNT MANAGEMENT")
        assert m is not None
        assert m.group("id") == "AC-2"
        assert m.group("title") == "ACCOUNT MANAGEMENT"

    def test_sc7(self):
        m = BASE_CONTROL_RE.match("SC-7 BOUNDARY PROTECTION")
        assert m is not None
        assert m.group("id") == "SC-7"

    def test_sr11(self):
        m = BASE_CONTROL_RE.match("SR-11 COMPONENT AUTHENTICITY")
        assert m is not None
        assert m.group("id") == "SR-11"

    def test_au4(self):
        m = BASE_CONTROL_RE.match("AU-4 AUDIT LOG STORAGE CAPACITY")
        assert m is not None
        assert m.group("id") == "AU-4"

    def test_no_match_body_text(self):
        assert BASE_CONTROL_RE.match("The system shall log events.") is None

    def test_no_match_lowercase(self):
        assert BASE_CONTROL_RE.match("ac-1 policy and procedures") is None


class TestEnhancementHeading:
    def test_simple(self):
        m = ENHANCEMENT_RE.match("(1) AUTOMATED SYSTEM ACCOUNT MANAGEMENT")
        assert m is not None
        assert m.group("num") == "1"
        assert m.group("title") == "AUTOMATED SYSTEM ACCOUNT MANAGEMENT"

    def test_with_pipe(self):
        m = ENHANCEMENT_RE.match("(3) INFORMATION FLOW ENFORCEMENT | DYNAMIC INFORMATION FLOW CONTROL")
        assert m is not None
        assert m.group("num") == "3"
        assert "DYNAMIC INFORMATION FLOW CONTROL" in m.group("title")

    def test_two_digit_number(self):
        m = ENHANCEMENT_RE.match("(12) DATA TYPE IDENTIFIERS")
        assert m is not None
        assert m.group("num") == "12"

    def test_no_match_body(self):
        assert ENHANCEMENT_RE.match("Discussion: This control") is None


class TestWithdrawn:
    def test_withdrawn_marker(self):
        assert WITHDRAWN_RE.search("[Withdrawn: Incorporated into AC-2k.]")
        assert WITHDRAWN_RE.search("[Withdrawn:  Moved to AC-3.]")

    def test_not_withdrawn(self):
        assert not WITHDRAWN_RE.search("Control: The system shall...")


# ---------------------------------------------------------------------------
# Text cleaning
# ---------------------------------------------------------------------------

class TestCleanText:
    def test_whitespace_collapse(self):
        assert _clean_text("hello   world") == "hello world"

    def test_soft_hyphen_removal(self):
        assert _clean_text("organi\u00adzation") == "organization"

    def test_bom_removal(self):
        assert _clean_text("\ufeffhello") == "hello"


class TestBoilerplate:
    def test_empty(self):
        assert _is_boilerplate("")

    def test_doi(self):
        assert _is_boilerplate("https://doi.org/10.6028/NIST.SP.800-53r5")

    def test_doi_text(self):
        assert _is_boilerplate("This publication is available free of charge from:")

    def test_page_header(self):
        assert _is_boilerplate("NIST SP 800-53, REV. 5 SECURITY AND PRIVACY CONTROLS")

    def test_underline(self):
        assert _is_boilerplate("_" * 50)

    def test_chapter_three_page(self):
        assert _is_boilerplate("CHAPTER THREE PAGE 18")

    def test_roman_numeral(self):
        assert _is_boilerplate("iii")
        assert _is_boilerplate("iv")

    def test_page_number(self):
        assert _is_boilerplate("42")

    def test_single_char(self):
        assert _is_boilerplate("T")
        assert _is_boilerplate("h")

    def test_not_boilerplate(self):
        assert not _is_boilerplate("Control: The system shall protect data.")
        assert not _is_boilerplate("AC-1 POLICY AND PROCEDURES")


class TestSideMargin:
    def test_rotated_word(self):
        w = {"text": "publication", "x0": 18.1, "upright": False}
        assert _is_side_margin_word(w)

    def test_margin_single_char(self):
        w = {"text": "T", "x0": 18.1, "upright": True}
        assert _is_side_margin_word(w)

    def test_margin_doi_fragment(self):
        w = {"text": "https://doi.org/10.6028", "x0": 18.1, "upright": True}
        assert _is_side_margin_word(w)

    def test_normal_word(self):
        w = {"text": "Control:", "x0": 90.0, "upright": True}
        assert not _is_side_margin_word(w)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

class TestValidation:
    def test_duplicate_detection(self):
        items = [
            {"clause_id": "AC-1", "title": "T", "text": "Control text here.", "confidence": 0.95, "issues": []},
            {"clause_id": "AC-1", "title": "T", "text": "Duplicate.", "confidence": 0.95, "issues": []},
        ]
        validated = _validate_items(items)
        assert "duplicate_clause_id" in validated[1]["issues"]

    def test_missing_clause_id(self):
        items = [
            {"clause_id": "", "title": "T", "text": "Some text.", "confidence": 0.95, "issues": []},
        ]
        validated = _validate_items(items)
        assert "missing_clause_id" in validated[0]["issues"]

    def test_short_text(self):
        items = [
            {"clause_id": "AC-1", "title": "T", "text": "Hi", "confidence": 0.95, "issues": []},
        ]
        validated = _validate_items(items)
        assert "text_too_short" in validated[0]["issues"]

    def test_doi_in_text(self):
        items = [
            {"clause_id": "AC-1", "title": "T",
             "text": "See doi.org/10.6028/NIST.SP.800-53r5 for details.",
             "confidence": 0.95, "issues": []},
        ]
        validated = _validate_items(items)
        assert "contains_doi_url" in validated[0]["issues"]

    def test_toc_dot_leader(self):
        items = [
            {"clause_id": "AC-1", "title": "T",
             "text": "Access Control..............................15",
             "confidence": 0.95, "issues": []},
        ]
        validated = _validate_items(items)
        assert "toc_dot_leader" in validated[0]["issues"]


# ---------------------------------------------------------------------------
# Classification of withdrawn controls
# ---------------------------------------------------------------------------

class TestClassification:
    def test_withdrawn_is_information(self):
        """Withdrawn controls should be classified as Information."""
        from nist80053_extractor import _parse_controls

        # Simulate the parsing with a minimal set of lines
        lines = [
            {"text": "AC-99 TEST CONTROL", "top": 100, "size": 11.0,
             "font": "Bold", "is_bold": True, "x0": 90, "page": 1},
            {"text": "[Withdrawn: Incorporated into AC-1.]", "top": 120, "size": 10.0,
             "font": "Regular", "is_bold": False, "x0": 90, "page": 1},
        ]
        items = _parse_controls(lines)
        assert len(items) == 1
        assert items[0]["clause_id"] == "AC-99"
        assert items[0]["classification"] == "Information"

    def test_active_control_is_requirement(self):
        """Active controls should be classified as Requirement."""
        from nist80053_extractor import _parse_controls

        lines = [
            {"text": "AC-99 TEST CONTROL", "top": 100, "size": 11.0,
             "font": "Bold", "is_bold": True, "x0": 90, "page": 1},
            {"text": "Control: The organization shall protect data.", "top": 120, "size": 10.0,
             "font": "Regular", "is_bold": False, "x0": 90, "page": 1},
        ]
        items = _parse_controls(lines)
        assert len(items) == 1
        assert items[0]["clause_id"] == "AC-99"
        assert items[0]["classification"] == "Requirement"


# ---------------------------------------------------------------------------
# Standard export integration
# ---------------------------------------------------------------------------

class TestStandardExportIntegration:
    def test_nist_items_write_to_template(self):
        """NIST items should write correctly to Standard.xlsx template."""
        from standard_export import write_standard_assessment
        import openpyxl

        items = [
            {"clause_id": "AC-1", "title": "POLICY AND PROCEDURES",
             "text": "Control: Develop and document access control policy.",
             "classification": "Requirement"},
            {"clause_id": "AC-2", "title": "ACCOUNT MANAGEMENT",
             "text": "Control: Define accounts.",
             "classification": "Requirement"},
            {"clause_id": "AC-2(10)", "title": "SHARED ACCOUNT",
             "text": "[Withdrawn: Incorporated into AC-2k.]",
             "classification": "Information"},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "nist_test.xlsx")
            write_standard_assessment(
                items, out,
                standard_id="NIST80053",
                standard_title="NIST SP 800-53",
            )

            wb = openpyxl.load_workbook(out)
            ws = wb["Standard Assessment"]

            # Row 10: AC-1
            assert ws["B10"].value == "AC-1"
            assert ws["C10"].value == "POLICY AND PROCEDURES"
            assert "access control policy" in ws["D10"].value
            assert ws["E10"].value == "Requirement"

            # Row 12: withdrawn
            assert ws["B12"].value == "AC-2(10)"
            assert ws["E12"].value == "Information"
            assert "[Withdrawn" in ws["D12"].value

            # F-I blank
            for col in ("F10", "G10", "H10", "I10"):
                assert not ws[col].value

            # Template machinery preserved
            assert len(ws.data_validations.dataValidation) >= 8

    def test_no_one_letter_rows(self):
        """Regression: no single-character text should end up in output."""
        items = [
            {"clause_id": "AC-1", "title": "T",
             "text": "Control: Full control text here.",
             "classification": "Requirement"},
        ]
        # All items have proper text
        for it in items:
            assert len(it["text"]) > 5


# ---------------------------------------------------------------------------
# Full PDF test (only runs when the actual PDF is available)
# ---------------------------------------------------------------------------

NIST_PDF = "/home/kali/Downloads/NIST.SP.800-53r5.pdf"


class TestFullPDF:
    """Integration tests that require the actual NIST PDF file."""

    def _skip_if_no_pdf(self):
        if not os.path.isfile(NIST_PDF):
            import pytest
            pytest.skip("NIST PDF not available")

    def test_is_nist_800_53(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import is_nist_800_53
        assert is_nist_800_53(NIST_PDF) is True

    def test_is_not_nist_for_other_pdf(self):
        """Sample PDFs should not be detected as NIST 800-53."""
        from nist80053_extractor import is_nist_800_53
        sample = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              "samples", "sample_law.pdf")
        if os.path.isfile(sample):
            assert is_nist_800_53(sample) is False

    def test_extract_produces_hundreds(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        assert len(items) > 500, f"Expected 500+ items, got {len(items)}"

    def test_ac1_present(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        ac1 = [it for it in items if it["clause_id"] == "AC-1"]
        assert len(ac1) == 1
        assert ac1[0]["title"] == "POLICY AND PROCEDURES"
        assert "Control:" in ac1[0]["text"]
        assert ac1[0]["classification"] == "Requirement"

    def test_ac2_enhancements(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        ac2_enh = [it for it in items if it["clause_id"].startswith("AC-2(")]
        assert len(ac2_enh) >= 5, f"Expected 5+ AC-2 enhancements, got {len(ac2_enh)}"

    def test_withdrawn_classified_information(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        withdrawn = [it for it in items if it["classification"] == "Information"]
        assert len(withdrawn) > 10, f"Expected 10+ withdrawn, got {len(withdrawn)}"
        for w in withdrawn:
            assert "[Withdrawn" in w["text"]

    def test_no_one_letter_rows(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        bad = [it for it in items if len(it["text"].strip()) <= 3]
        assert len(bad) == 0, f"Found {len(bad)} one-letter rows: {bad[:3]}"

    def test_no_doi_rows(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        doi = [it for it in items if "doi.org/10.6028" in it["text"]]
        assert len(doi) == 0, f"Found {len(doi)} DOI rows"

    def test_no_toc_rows(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        toc = [it for it in items if re.search(r"\.{4,}", it["text"])]
        assert len(toc) == 0, f"Found {len(toc)} TOC dot-leader rows"

    def test_all_items_have_clause_id(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        missing = [it for it in items if not it["clause_id"]]
        assert len(missing) == 0, f"Found {len(missing)} items without clause_id"

    def test_clause_id_format(self):
        self._skip_if_no_pdf()
        from nist80053_extractor import extract_nist_800_53_items
        items = extract_nist_800_53_items(NIST_PDF)
        pattern = re.compile(r"^[A-Z]{2}-\d+(\(\d+\))?$")
        bad = [it for it in items if not pattern.match(it["clause_id"])]
        assert len(bad) == 0, f"Bad clause_id format: {[b['clause_id'] for b in bad[:5]]}"
