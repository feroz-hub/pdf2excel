"""Offline test: enriched items write cols F–I and preserve the template.

Run with pytest (``pytest tests/``) or directly
(``python tests/test_standard_export_enriched.py``).
"""

import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl                          # noqa: E402
from standard_export import write_standard_assessment  # noqa: E402


def _enriched_items():
    return [
        {"clause_id": "Article 4", "title": "Obligations",
         "text": "Operators shall keep records.", "classification": "Requirement",
         "requirement": "The organization shall retain records for five years.",
         "detailed_description": "REQ-001 | Verification: Audit | Direct\nRetain "
                                 "and produce records on request.",
         "change_in_requirement": "Establish a 5-year retention store.",
         "requirement_classification": "Process"},
        {"clause_id": "Article 1", "title": "Purpose",
         "text": "Defines scope.", "classification": "Information",
         "requirement": "", "detailed_description": "",
         "change_in_requirement": "", "requirement_classification": ""},
    ]


def test_enriched_write_fills_F_to_I_and_preserves_template():
    out = os.path.join(tempfile.mkdtemp(prefix="pdf2excel_test_"), "enriched.xlsx")
    write_standard_assessment(_enriched_items(), out, standard_id="MLSR999")

    wb = openpyxl.load_workbook(out)
    ws = wb["Standard Assessment"]

    # Requirement row (10): E–I populated.
    assert ws["E10"].value == "Requirement"
    assert ws["F10"].value.startswith("The organization shall retain")
    assert ws["G10"].value.startswith("REQ-001 | Verification: Audit")
    assert ws["H10"].value == "Establish a 5-year retention store."
    assert ws["I10"].value == "Process"

    # Information row (11): F–I blank.
    assert ws["E11"].value == "Information"
    for col in ("F11", "G11", "H11", "I11"):
        assert not ws[col].value, f"{col} should be blank, got {ws[col].value!r}"

    # Analyst columns J–Q untouched on data rows.
    assert all(ws.cell(row=10, column=c).value is None for c in range(10, 18))

    # Template machinery survives (rows cleared by value, never deleted).
    assert len(ws.data_validations.dataValidation) >= 8
    dn = wb.defined_names
    names = list(dn.keys()) if hasattr(dn, "keys") else list(dn)
    for nm in ("CLASSIFICATION", "Requirement_Applicability", "Responsible_by"):
        assert nm in names, f"named range {nm} lost"
    assert "O8:Q8" in [str(m) for m in ws.merged_cells.ranges]


def test_non_enriched_items_leave_F_to_I_blank():
    # Backward compatibility: items without the E–I keys behave as before.
    items = [{"clause_id": "A1", "title": "Purpose", "text": "Scope.",
              "classification": "Information"}]
    out = os.path.join(tempfile.mkdtemp(prefix="pdf2excel_test_"), "base.xlsx")
    write_standard_assessment(items, out, standard_id="MLSR1")
    ws = openpyxl.load_workbook(out)["Standard Assessment"]
    assert ws["D10"].value == "Scope." and ws["E10"].value == "Information"
    for col in ("F10", "G10", "H10", "I10"):
        assert not ws[col].value


def _run_all():
    fns = [v for k, v in sorted(globals().items())
           if k.startswith("test_") and callable(v)]
    for fn in fns:
        fn()
        print(f"  PASS {fn.__name__}")
    print(f"standard_export(enriched): {len(fns)} tests passed")


if __name__ == "__main__":
    _run_all()
