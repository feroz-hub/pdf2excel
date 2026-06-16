"""Export extractor output into the "Standard Assessment" Excel format.

This writes into ``templates/Standard.xlsx`` **as a template** — it is loaded
with openpyxl and only the metadata cells and the data rows (10+) are touched.
Merged cells, the Times New Roman header styling, the data-validation dropdowns
and the named ranges (CLASSIFICATION / Requirement_Applicability /
Responsible_by / Information) are left untouched so they keep working.

The data rows are *cleared by value* (set to ``None``) rather than deleted —
deleting rows would corrupt the data validations whose ranges (E10:E1020,
I10:I1020, ...) are anchored to specific rows.

Public API:
    write_standard_assessment(items, out_path, template_path=..., standard_id=...,
                              standard_title=..., standard_edition=...,
                              document_id=..., document_name=..., document_revision=...)
    paragraphs_to_items(paras)         # prose -> items
    deck_to_items(pages_tables, slides_text)  # deck -> items

An ``item`` is a dict: {clause_id, title, text, classification}.
"""

from __future__ import annotations

import os
import re
from typing import List

# Default location of the bundled template, resolved relative to this file so
# it works regardless of the current working directory.
DEFAULT_TEMPLATE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "templates", "Standard.xlsx"
)

SHEET_NAME = "Standard Assessment"
DATA_START_ROW = 10
_LAST_COL = 17  # column Q

# Leading legal clause token, e.g. "Article 1", "Chapter 2", "Annex 3".
_HEADING_ID_RE = re.compile(
    r"^(Article|Chapter|Section|Part|Annex|Schedule)\s+\d+", re.IGNORECASE
)
# Fallback for numbered headings, e.g. "4.2 Documentation".
_NUMBERED_ID_RE = re.compile(r"^(\d+(?:\.\d+)*)\.?\s+(.*)$")
# Obligation language that promotes a paragraph to a "Requirement".
_REQUIREMENT_RE = re.compile(
    r"\b(shall not|shall|must|is required to|are required to)\b", re.IGNORECASE
)


# --------------------------------------------------------------------------- #
# Mapping layer: extractor output -> items
# --------------------------------------------------------------------------- #

def split_heading(label: str):
    """Split a heading label into (clause_id, title).

    "Article 1 (Purpose)"       -> ("Article 1", "Purpose")
    "Chapter 2 General Matters" -> ("Chapter 2", "General Matters")
    "4.2 Documentation"         -> ("4.2", "Documentation")
    Anything unrecognised        -> (label, "").
    """
    label = (label or "").strip()
    m = _HEADING_ID_RE.match(label)
    if m:
        clause_id = m.group(0).strip()
        rest = label[m.end():].strip()
        if rest.startswith("(") and rest.endswith(")"):
            rest = rest[1:-1].strip()
        return clause_id, rest
    m = _NUMBERED_ID_RE.match(label)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return label, ""


def _classify(text: str, classify_requirements: bool) -> str:
    if classify_requirements and _REQUIREMENT_RE.search(text or ""):
        return "Requirement"
    return "Information"


def paragraphs_to_items(paras, classify_requirements: bool = True) -> List[dict]:
    """Turn prose paragraphs into items.

    Heading paragraphs update the running clause id/title but emit no row of
    their own; each body paragraph becomes one item.
    """
    items: List[dict] = []
    cur_id, cur_title = "", ""
    for p in paras:
        if getattr(p, "type", "body") == "heading":
            cur_id, cur_title = split_heading(p.text)
            continue
        items.append(
            {
                "clause_id": cur_id,
                "title": cur_title,
                "text": p.text,
                "classification": _classify(p.text, classify_requirements),
            }
        )
    return items


def _slide_title(text: str, page_no: int) -> str:
    for line in (text or "").splitlines():
        if line.strip():
            return line.strip()
    return f"Slide {page_no}"


def deck_to_items(pages_tables, slides_text) -> List[dict]:
    """Turn deck output into items.

    Every table row becomes one item (cells joined by " | "), and every slide's
    full text becomes one item, so nothing is lost.
    """
    title_by_page = {}
    text_by_page = {}
    for page_no, text in slides_text:
        title_by_page[page_no] = _slide_title(text, page_no)
        text_by_page[page_no] = (text or "").strip()

    tables_by_page = {}
    for page_no, tables in pages_tables:
        tables_by_page.setdefault(page_no, []).extend(tables)

    pages = sorted(set(title_by_page) | set(tables_by_page))
    items: List[dict] = []
    for page_no in pages:
        title = title_by_page.get(page_no, f"Slide {page_no}")
        for idx, table in enumerate(tables_by_page.get(page_no, [])):
            clause = f"slide{page_no}" if idx == 0 else f"slide{page_no}_{idx + 1}"
            for row in table.rows:
                items.append(
                    {
                        "clause_id": clause,
                        "title": title,
                        "text": " | ".join(c for c in row),
                        "classification": "Information",
                    }
                )
        if page_no in text_by_page:
            items.append(
                {
                    "clause_id": f"slide{page_no}",
                    "title": title,
                    "text": text_by_page[page_no],
                    "classification": "Information",
                }
            )
    return items


def blocks_to_items(blocks: List[dict], classify_requirements: bool = True) -> List[dict]:
    """Convert extracted blocks into Standard Assessment rows.

    Headings update current clause id and title; body paragraphs/notes/table rows
    become individual rows; TOC blocks are skipped by default.
    """
    items: List[dict] = []
    cur_id, cur_title = "", ""
    for b in blocks:
        b_type = b.get("block_type", "paragraph")
        
        if b_type == "heading":
            cur_id, cur_title = split_heading(b["text"])
            continue
            
        if b_type == "toc":
            continue
            
        if b_type == "table":
            rows = b.get("rows")
            if rows:
                for row in rows:
                    items.append({
                        "page": b.get("page"),
                        "clause_id": cur_id,
                        "title": cur_title,
                        "text": " | ".join(str(c or "").strip() for c in row),
                        "classification": "Information",
                        "confidence": b.get("confidence", 0.9),
                        "issues": list(b.get("issues", [])),
                        "bbox": b.get("bbox")
                    })
            else:
                for row_text in b["text"].splitlines():
                    if row_text.strip():
                        items.append({
                            "page": b.get("page"),
                            "clause_id": cur_id,
                            "title": cur_title,
                            "text": row_text.strip(),
                            "classification": "Information",
                            "confidence": b.get("confidence", 0.9),
                            "issues": list(b.get("issues", [])),
                            "bbox": b.get("bbox")
                        })
        elif b_type == "note":
            items.append({
                "page": b.get("page"),
                "clause_id": cur_id,
                "title": cur_title,
                "text": b["text"],
                "classification": "Information",
                "confidence": b.get("confidence", 1.0),
                "issues": list(b.get("issues", [])),
                "bbox": b.get("bbox")
            })
        else:  # paragraph, unknown
            classification = _classify(b["text"], classify_requirements)
            items.append({
                "page": b.get("page"),
                "clause_id": cur_id,
                "title": cur_title,
                "text": b["text"],
                "classification": classification,
                "confidence": b.get("confidence", 1.0),
                "issues": list(b.get("issues", [])),
                "bbox": b.get("bbox")
            })
    return items


# --------------------------------------------------------------------------- #
# Writer
# --------------------------------------------------------------------------- #

def write_standard_assessment(
    items,
    out_path: str,
    template_path: str = DEFAULT_TEMPLATE,
    standard_id: str = "MLSR",
    standard_title: str = "",
    standard_edition: str = "",
    document_id: str = "",
    document_name: str = "",
    document_revision: str = "",
    export_items=None,
    show_issues: bool = True,
) -> None:
    """Populate the Standard Assessment template and save it to ``out_path``."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"template not found: {template_path}")

    wb = load_workbook(template_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"template has no '{SHEET_NAME}' sheet")
    ws = wb[SHEET_NAME]

    # Metadata block: overwrite the external-link formulas with plain values
    # (blank when not provided). These cells are the masters of merged ranges.
    ws["B1"] = document_id or None
    ws["B2"] = document_name or None
    ws["B3"] = document_revision or None
    ws["B5"] = standard_id or None
    ws["B6"] = standard_title or None
    ws["B7"] = standard_edition or None

    # Clear existing data-row VALUES (do NOT delete rows: that breaks the
    # validations anchored to E10:E1020 / I10:I1020 / ...).
    last_row = ws.max_row
    for r in range(DATA_START_ROW, last_row + 1):
        for c in range(1, _LAST_COL + 1):
            ws.cell(row=r, column=c).value = None

    font = Font(name="Times New Roman", size=10)
    align = Alignment(vertical="top", wrap_text=True)

    actual_export = export_items if export_items is not None else items

    prev_clause = None
    for i, item in enumerate(actual_export, start=1):
        row = DATA_START_ROW + i - 1
        clause = (item.get("clause_id") or "").strip()
        values = {
            1: f"{standard_id}_{i}",
            # B is sparse: written only when the clause changes (like the template).
            2: clause if clause != prev_clause else None,
            3: item.get("title", ""),
            4: item.get("text", ""),
            5: item.get("classification", "Information") or "Information",
            # F–I: AI-enrichment output (ai_enrich). Absent keys write "", so
            # non-enriched items behave exactly as before (F–I left blank).
            6: item.get("requirement", "") or "",
            7: item.get("detailed_description", "") or "",
            8: item.get("change_in_requirement", "") or "",
            9: item.get("requirement_classification", "") or "",
            # J..Q intentionally left empty for the analyst (dropdowns remain).
        }
        for c, v in values.items():
            cell = ws.cell(row=row, column=c)
            cell.value = v
            cell.font = font
            cell.alignment = align
        prev_clause = clause

    # Write Extraction_Issues sheet when requested and issues exist
    has_issues = show_issues and any(it.get("issues") for it in items)
    if has_issues:
        if "Extraction_Issues" in wb.sheetnames:
            ws_issues = wb["Extraction_Issues"]
            # clear sheet
            ws_issues.delete_rows(1, ws_issues.max_row)
        else:
            ws_issues = wb.create_sheet(title="Extraction_Issues")
            
        headers = ["Row", "Page", "Clause ID", "Title", "Text", "Confidence", "Issues"]
        ws_issues.append(headers)
        
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="C00000")
        body_font = Font(name="Arial", size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")
        top_align = Alignment(vertical="top")
        
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_issues.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for i, item in enumerate(items, start=1):
            issues = item.get("issues", [])
            if issues:
                row_num = DATA_START_ROW + i - 1
                ws_issues.append([
                    row_num,
                    item.get("page", ""),
                    item.get("clause_id", ""),
                    item.get("title", ""),
                    item.get("text", ""),
                    item.get("confidence", 1.0),
                    ", ".join(issues)
                ])
                r_idx = ws_issues.max_row
                for col_idx in range(1, len(headers) + 1):
                    cell = ws_issues.cell(row=r_idx, column=col_idx)
                    cell.font = body_font
                    cell.alignment = wrap_align if col_idx in (4, 5) else top_align
                    
        widths = {1: 8, 2: 8, 3: 12, 4: 20, 5: 50, 6: 12, 7: 30}
        for col_idx, w in widths.items():
            ws_issues.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(out_path)

