"""Block-based extraction layer for pdf2excel.

Segments page content into unified blocks: headings, paragraphs, tables, notes, toc,
and orders them by page and y-position (handling multi-column layouts).
"""

from __future__ import annotations

import re
import logging
import statistics
from typing import Any, Dict, List, Optional, Tuple

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from deck_tables_to_excel import _extract_page_tables, _merge_side_by_side
from pdf_paragraphs_to_excel import _clean_text, _looks_like_page_number, _SENTENCE_END_RE, _HYPHEN_END_RE, _normalize_digits

log = logging.getLogger("pdf2excel.extract_blocks")

# Heading patterns to match standard headings
HEADING_PATTERNS = [
    # 1 Scope, 1.1 General, 1.1.1 Requirement, 4.2 Documentation, 4.2.1.2.3 etc.
    re.compile(r"^\d+(\.\d+){0,6}\.?\s+\S"),
    # Article 1, Chapter 2, Section 3, Annex A, Appendix B, Control 5.1, Requirement 1, Table 1
    re.compile(r"^(article|chapter|section|part|annex|appendix|control|requirement|table)\s+[a-z\d]+(\.[a-z\d]+)*\b", re.IGNORECASE),
    # A.1 General, B.2.3 etc.
    re.compile(r"^[a-z](\.\d+)+\s+\S", re.IGNORECASE),
    # NOTE 1 (if treated as heading structure)
    re.compile(r"^(note)\s+\d+\b", re.IGNORECASE)
]

LIST_START_RE = re.compile(r"^([•\-\*\▪\◦\♦]|\([a-z0-9]+\)|[a-z0-9]+\))\s+", re.IGNORECASE)
NOTE_START_RE = re.compile(r"^(note|remark|caution|warning)\b", re.IGNORECASE)


class BlockLine:
    """A visual line of text on a page with layout metrics."""
    def __init__(self, text: str, top: float, bottom: float, x0: float, x1: float, font_size: float, is_bold: bool, words: List[dict]):
        self.text = text
        self.top = top
        self.bottom = bottom
        self.x0 = x0
        self.x1 = x1
        self.font_size = font_size
        self.is_bold = is_bold
        self.words = words


def extract_page_tables_with_fallbacks(page: Any) -> List[Any]:
    """Extract tables using standard pdfplumber methods and text-aligned borderless fallbacks."""
    # 1. Standard tables
    tables = _extract_page_tables(page)
    
    # 2. Text-aligned borderless tables fallback if none found
    if not tables:
        try:
            from deck_tables_to_excel import _Table, _clean_cell, _is_real_table
            # Find borderless tables using text strategy
            for tbl in page.find_tables(table_settings={
                "vertical_strategy": "text",
                "horizontal_strategy": "text",
                "snap_y_tolerance": 5,
                "snap_x_tolerance": 5,
            }):
                raw = tbl.extract()
                rows = [[_clean_cell(c) for c in row] for row in raw]
                if _is_real_table(rows):
                    tables.append(_Table(rows=rows, bbox=tuple(tbl.bbox)))
        except Exception as exc:
            log.warning("Borderless table fallback failed: %s", exc)
            
    return _merge_side_by_side(tables)


def cluster_prose_words(words: List[dict]) -> List[BlockLine]:
    """Group pdfplumber word boxes into visual lines, extracting font size and boldness."""
    if not words:
        return []
        
    # Sort top-to-bottom, then left-to-right
    words = sorted(words, key=lambda w: (round(w["top"], 1), w["x0"]))
    
    lines: List[List[dict]] = []
    current = [words[0]]
    current_top = words[0]["top"]
    
    for w in words[1:]:
        height = max(w["bottom"] - w["top"], 1.0)
        tol = height * 0.6
        if abs(w["top"] - current_top) <= tol:
            current.append(w)
        else:
            lines.append(current)
            current = [w]
            current_top = w["top"]
    lines.append(current)
    
    result = []
    for group in lines:
        group = sorted(group, key=lambda w: w["x0"])
        text = " ".join(w["text"] for w in group)
        text = _clean_text(text)
        if not text:
            continue
            
        x0 = min(w["x0"] for w in group)
        x1 = max(w["x1"] for w in group)
        top = min(w["top"] for w in group)
        bottom = max(w["bottom"] for w in group)
        
        sizes = [w.get("size") for w in group if w.get("size") is not None]
        avg_size = sum(sizes) / len(sizes) if sizes else 10.0
        
        is_bold = False
        for w in group:
            fn = str(w.get("fontname", "")).lower()
            if any(k in fn for k in ("bold", "black", "heavy", "semibold")):
                is_bold = True
                break
                
        result.append(BlockLine(
            text=text, top=top, bottom=bottom, x0=x0, x1=x1,
            font_size=avg_size, is_bold=is_bold, words=group
        ))
    return result


def is_heading(line: BlockLine, doc_median_size: float, median_gap: float, next_line_gap: float) -> bool:
    """Determine if a visual line is a heading based on pattern and visual layout clues."""
    text = line.text.strip()
    
    # 1. Matches regular heading patterns (Article 1, 1.1 General)
    for p in HEADING_PATTERNS:
        if p.match(text) and len(text) < 150:
            return True
            
    # 2. Significantly larger font size than document median body text
    if line.font_size >= doc_median_size + 1.5 and len(text) < 120:
        return True
        
    # 3. Bold font, short line, starts near left margin, followed by a gap
    if line.is_bold and len(text) < 80 and next_line_gap > median_gap * 1.4:
        return True
        
    return False


def join_line_texts(lines: List[BlockLine]) -> str:
    """Join line text with hyphen removal."""
    out = ""
    for ln in lines:
        piece = ln.text
        if not out:
            out = piece
            continue
        m = _HYPHEN_END_RE.search(out)
        if m:
            out = out[:m.start(1) + 1] + piece.lstrip()
        else:
            out = out + " " + piece
    return _clean_text(out)


def group_lines_into_blocks(
    lines: List[BlockLine],
    page_num: int,
    page_type: str,
    gap_factor: float,
    doc_median_size: float
) -> List[dict]:
    """Group visual lines into paragraph, heading, note, or TOC blocks using layout cues."""
    if not lines:
        return []
        
    # Compute median gap between lines
    gaps = []
    for prev, cur in zip(lines, lines[1:]):
        gap = cur.top - prev.bottom
        if gap > 0:
            gaps.append(gap)
    median_gap = statistics.median(gaps) if gaps else 0.0
    gap_threshold = median_gap * gap_factor if median_gap > 0 else 0.0
    
    blocks = []
    buffer: List[BlockLine] = []
    current_type = "paragraph"
    
    def flush():
        if not buffer:
            return
        text = join_line_texts(buffer)
        if text:
            # Determine overall block boundaries
            x0 = min(ln.x0 for ln in buffer)
            top = min(ln.top for ln in buffer)
            x1 = max(ln.x1 for ln in buffer)
            bottom = max(ln.bottom for ln in buffer)
            
            # Map type if the page is TOC
            block_type = "toc" if page_type == "toc" else current_type
            
            blocks.append({
                "page": page_num,
                "block_type": block_type,
                "clause_id": "",
                "title": "",
                "text": text,
                "bbox": [x0, top, x1, bottom],
                "confidence": 1.0,
                "issues": []
            })
        buffer.clear()
        
    for idx, ln in enumerate(lines):
        # Determine gap to next line for heading check
        next_gap = 0.0
        if idx + 1 < len(lines):
            next_gap = lines[idx + 1].top - ln.bottom
            
        # 1. Heading check
        if is_heading(ln, doc_median_size, median_gap, next_gap):
            flush()
            # Headings are single-line blocks
            buffer.append(ln)
            current_type = "heading"
            flush()
            current_type = "paragraph"
            continue
            
        # 2. Note check
        if NOTE_START_RE.match(ln.text):
            flush()
            buffer.append(ln)
            current_type = "note"
            continue
            
        # 3. Standard line: decide whether to group or break
        if buffer:
            prev = buffer[-1]
            gap = ln.top - prev.bottom
            
            # Decide break signals
            is_break = False
            if gap_threshold and gap > gap_threshold:
                is_break = True
            elif abs(ln.font_size - prev.font_size) > 1.2:
                is_break = True
            elif ln.is_bold != prev.is_bold:
                is_break = True
            elif abs(ln.x0 - prev.x0) > 15:
                is_break = True
            elif LIST_START_RE.match(ln.text):
                is_break = True
            elif _SENTENCE_END_RE.search(prev.text) and ln.text[:1].isupper():
                is_break = True
                
            if is_break:
                flush()
                if NOTE_START_RE.match(ln.text):
                    current_type = "note"
                else:
                    current_type = "paragraph"
                    
        buffer.append(ln)
        
    flush()
    return blocks


def sort_blocks_layout(blocks: List[dict], page_width: float, is_multi_column: bool) -> List[dict]:
    """Sort blocks by page reading order.
    
    If multi-column, sorts column blocks left-to-right and top-to-bottom,
    preserving full-width layout dividers (like headers, spans).
    """
    if not is_multi_column:
        return sorted(blocks, key=lambda b: b["bbox"][1])
        
    # Sort top-to-bottom first to process bands
    sorted_by_y = sorted(blocks, key=lambda b: b["bbox"][1])
    center_x = page_width / 2
    
    bands = []
    current_band = []
    
    for b in sorted_by_y:
        x0, top, x1, bottom = b["bbox"]
        width = x1 - x0
        # A block is full-width if it spans >60% of page and crosses the center line
        is_full_width = (width > page_width * 0.6) and (x0 < center_x < x1)
        
        if is_full_width:
            if current_band:
                bands.append((False, current_band))
                current_band = []
            bands.append((True, [b]))
        else:
            current_band.append(b)
            
    if current_band:
        bands.append((False, current_band))
        
    final_blocks = []
    for is_fw, band_blocks in bands:
        if is_fw:
            final_blocks.extend(band_blocks)
        else:
            left_col = []
            right_col = []
            for b in band_blocks:
                x0, top, x1, bottom = b["bbox"]
                cx = (x0 + x1) / 2
                if cx < center_x:
                    left_col.append(b)
                else:
                    right_col.append(b)
            final_blocks.extend(sorted(left_col, key=lambda b: b["bbox"][1]))
            final_blocks.extend(sorted(right_col, key=lambda b: b["bbox"][1]))
            
    return final_blocks


def write_review_workbook(blocks: List[dict], out_path: str) -> None:
    """Write intermediate review workbook containing raw blocks."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Blocks"
    
    headers = ["page", "block_id", "block_type", "clause_id", "title", "text", "confidence", "issues", "bbox"]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")
    
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")
        
    for b in blocks:
        ws.append([
            b["page"],
            b.get("block_id", ""),
            b["block_type"],
            b.get("clause_id", ""),
            b.get("title", ""),
            b["text"],
            b["confidence"],
            ", ".join(b.get("issues", [])),
            str(b["bbox"])
        ])
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.alignment = wrap if col in (5, 6) else top_align
            
    widths = {1: 8, 2: 10, 3: 12, 4: 12, 5: 20, 6: 60, 7: 12, 8: 20, 9: 25}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
        
    ws.freeze_panes = "A2"
    wb.save(out_path)


def extract_page_blocks(
    page: Any,
    preflight_meta: Dict[str, Any],
    gap_factor: float,
    doc_median_size: float,
    boilerplate_set: Set[str],
    ocr_mode: str = "off"
) -> List[Dict[str, Any]]:
    """Extract and group blocks for a single page based on preflight metadata."""
    page_num = preflight_meta["page_number"]
    page_type = preflight_meta["page_type"]
    
    # 1. Scanned page handling
    if preflight_meta["is_scanned"]:
        ocr_text = preflight_meta.get("ocr_text", "")
        if ocr_text:
            lines = []
            for line_str in ocr_text.splitlines():
                line_str = line_str.strip()
                if line_str:
                    lines.append(BlockLine(
                        text=line_str, top=0, bottom=0, x0=0, x1=page.width,
                        font_size=doc_median_size, is_bold=False, words=[]
                    ))
            return group_lines_into_blocks(lines, page_num, page_type, gap_factor, doc_median_size)
        else:
            return [{
                "page": page_num,
                "block_type": "unknown",
                "clause_id": "",
                "title": "",
                "text": "[Scanned Page — Selectable Text Missing]",
                "bbox": [0.0, 0.0, page.width, page.height],
                "confidence": 0.1,
                "issues": ["scanned_page_skipped"]
            }]
            
    # 2. Extract tables
    tables = []
    if page_type in ("table", "mixed", "unknown"):
        tables = extract_page_tables_with_fallbacks(page)
        
    # 3. Extract prose lines
    prose_lines = []
    if page_type in ("prose", "mixed", "cover", "toc", "unknown"):
        table_bboxes = [t.bbox for t in tables]
        words = page.extract_words(
            use_text_flow=False,
            keep_blank_chars=False,
            extra_attrs=["size", "fontname"]
        )
        
        # Filter words not in any table bbox
        prose_words = []
        for w in words:
            cx = (w["x0"] + w["x1"]) / 2
            cy = (w["top"] + w["bottom"]) / 2
            inside_table = False
            for bx0, btop, bx1, bbottom in table_bboxes:
                if bx0 - 2 <= cx <= bx1 + 2 and btop - 2 <= cy <= bbottom + 2:
                    inside_table = True
                    break
            if not inside_table:
                prose_words.append(w)
                
        raw_lines = cluster_prose_words(prose_words)
        
        # Filter boilerplate (headers/footers/page numbers)
        for ln in raw_lines:
            norm = _normalize_digits(ln.text).strip()
            if norm in boilerplate_set or _looks_like_page_number(ln.text):
                continue
            prose_lines.append(ln)
            
    # 4. Group lines into blocks
    blocks = group_lines_into_blocks(prose_lines, page_num, page_type, gap_factor, doc_median_size)
    
    # 5. Convert tables to blocks
    for tbl in tables:
        text_rows = []
        for row in tbl.rows:
            text_rows.append(" | ".join(cell or "" for cell in row))
        table_text = "\n".join(text_rows)
        
        blocks.append({
            "page": page_num,
            "block_type": "table",
            "clause_id": "",
            "title": "",
            "text": table_text,
            "bbox": list(tbl.bbox),
            "confidence": 0.9,
            "issues": [],
            "rows": tbl.rows
        })
        
    # 6. Sort by reading order
    is_multi_column = preflight_meta["is_multi_column"]
    return sort_blocks_layout(blocks, page.width, is_multi_column)


def extract_pdf_blocks(
    pdf_path: str,
    page_preflights: List[Dict[str, Any]],
    gap_factor: float = 1.6,
    ocr_mode: str = "off",
    skip_cover: bool = True,
    include_toc: bool = False
) -> List[Dict[str, Any]]:
    """Orchestrates block-level extraction across all pages of a PDF."""
    if not pdfplumber:
        raise RuntimeError("pdfplumber is not installed.")
        
    # Calculate document-wide median font size
    sizes = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for w in page.extract_words(extra_attrs=["size"]):
                if w.get("size") is not None:
                    sizes.append(w["size"])
    doc_median_size = statistics.median(sizes) if sizes else 10.0
    
    # Document-wide boilerplate detection
    # Step 1: Collect prose lines from each page (excluding skipped pages)
    all_pages_lines = []
    preflight_by_page = {p["page_number"]: p for p in page_preflights}
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            meta = preflight_by_page.get(page_num)
            if not meta:
                continue
            if skip_cover and meta["page_type"] == "cover":
                continue
            if not include_toc and meta["page_type"] == "toc":
                continue
            if meta["is_scanned"]:
                continue
                
            # Temporary extraction to detect boilerplate
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            # Find and ignore tables to get clean prose lines for boilerplate
            tables = _extract_page_tables(page)
            table_bboxes = [t.bbox for t in tables]
            prose_words = []
            for w in words:
                cx = (w["x0"] + w["x1"]) / 2
                cy = (w["top"] + w["bottom"]) / 2
                inside = False
                for bx0, btop, bx1, bbottom in table_bboxes:
                    if bx0 - 2 <= cx <= bx1 + 2 and btop - 2 <= cy <= bbottom + 2:
                        inside = True
                        break
                if not inside:
                    prose_words.append(w)
                    
            lines = cluster_prose_words(prose_words)
            # Adapt line types for the repeated header detector which expects _Line class
            # (which has text field). We can just pass the BlockLine objects.
            all_pages_lines.append(lines)
            
    # Detect repeating headers/footers
    from pdf_paragraphs_to_excel import _detect_repeating_headers_footers
    boilerplate_set = _detect_repeating_headers_footers(all_pages_lines)
    
    # Step 2: Perform the actual block extraction
    all_blocks = []
    block_id_counter = 1
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            meta = preflight_by_page.get(page_num)
            if not meta:
                continue
            if skip_cover and meta["page_type"] == "cover":
                log.info("Page %d: Skipping cover page.", page_num)
                continue
            if not include_toc and meta["page_type"] == "toc":
                log.info("Page %d: Skipping TOC page.", page_num)
                continue
                
            page_blocks = extract_page_blocks(
                page=page,
                preflight_meta=meta,
                gap_factor=gap_factor,
                doc_median_size=doc_median_size,
                boilerplate_set=boilerplate_set,
                ocr_mode=ocr_mode
            )
            
            for b in page_blocks:
                b["block_id"] = block_id_counter
                block_id_counter += 1
                all_blocks.append(b)
                
    return all_blocks

