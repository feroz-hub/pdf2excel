"""Extract paragraphs from a text-based PDF into an Excel workbook.

This module contains the extraction engine and a command-line interface.

The engine reconstructs paragraphs from the raw word boxes that ``pdfplumber``
reports, rather than relying on the PDF's own (often unreliable) text flow:

  1. Words are clustered into visual *lines* using their y-coordinates.
  2. Lines are grouped into *paragraphs* by looking for vertical gaps that are
     noticeably larger than the page's typical line spacing.
  3. Running headers / footers and page numbers are detected by their repeated
     (digit-normalized) text across pages, and stripped.
  4. Numbered section headings (e.g. ``4.2 Scope``) are detected and tracked so
     every output row carries its nearest heading.
  5. Words hyphenated across a line break are re-joined, and paragraphs that
     continue across a page boundary are merged.

Public API:
    extract_paragraphs(pdf_path, gap_factor=1.6) -> list[Paragraph]
    write_excel(paras, out_path) -> None
"""

from __future__ import annotations

import argparse
import re
import statistics
import sys
from dataclasses import dataclass, field
from typing import List, Optional

try:
    import pdfplumber
except ImportError:  # pragma: no cover - dependency hint
    pdfplumber = None


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #

@dataclass
class Paragraph:
    """One extracted paragraph destined for a spreadsheet row."""

    para_id: int = 0
    page: int = 0            # 1-based page number where the paragraph starts
    type: str = "body"       # "heading" or "body"
    section: str = ""        # nearest section heading text
    text: str = ""


@dataclass
class _Line:
    """An internal, intermediate visual line of text on a page."""

    text: str
    top: float               # y of the line's top (smaller = higher on page)
    bottom: float            # y of the line's bottom
    x0: float                # left edge of the line
    words: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Regexes / constants
# --------------------------------------------------------------------------- #

# Unmapped-glyph artifacts that pdfplumber emits for fonts it cannot decode.
_CID_RE = re.compile(r"\(cid:\d+\)")

# Heading styles, each a (compiled-regex, max_len) pair. A line is a heading of
# the style when the pattern matches AND the line is shorter than max_len (real
# headings are short; long matches are usually running prose or citations).
#
# numbered: "4", "4.2", "4.2.1 Documentation" — dotted clause numbers.
# legal:    "Article 7 (...)", "Chapter 2 ...", "Section 1" — the lookahead after
#           the number stops us from matching in-text citations such as
#           "Article 2, Paragraph 1 shall apply".
HEADING_STYLES = {
    "numbered": (re.compile(r"^\d+(\.\d+){0,4}\.?\s+\S"), 90),
    "legal": (
        re.compile(
            r"^(chapter|article|section|part|annex|schedule)\s+\d+"
            r"(?=\s*\(|\s+[A-Z]|\s*$)",
            re.IGNORECASE,
        ),
        130,
    ),
}

# Splits a parenthetical legal heading ("Article 1 (Purpose) The purpose ...")
# into its clean label and the trailing body sentence that follows it.
_HEADING_LABEL_RE = re.compile(
    r"^((?:article|annex|schedule|part)\s+\d+\s*\([^)]*\))\s*(.*)$",
    re.IGNORECASE,
)

# Sentence-ending punctuation used to decide whether a paragraph is "complete".
_SENTENCE_END_RE = re.compile(r"[.!?:;)\"”’]\s*$")

# Trailing hyphen indicating a word split across a line break.
_HYPHEN_END_RE = re.compile(r"(\w)[-‐]$")


# --------------------------------------------------------------------------- #
# Text helpers
# --------------------------------------------------------------------------- #

def _normalize_digits(text: str) -> str:
    """Replace every digit with '#' so "Page 1" and "Page 2" collapse together."""
    return re.sub(r"\d", "#", text)


def _clean_text(text: str) -> str:
    """Remove cid artifacts and collapse runaway whitespace."""
    text = _CID_RE.sub("", text)
    text = text.replace("­", "")          # soft hyphen
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def _looks_like_page_number(text: str) -> bool:
    """True for lines that are essentially just a page number."""
    stripped = text.strip()
    if not stripped:
        return False
    # "12", "Page 12", "12 / 30", "- 12 -", "Page 12 of 30"
    return bool(
        re.fullmatch(
            r"(page\s+)?[-–\s]*\d+([\s/]+(of\s+)?\d+)?[-–\s]*",
            stripped,
            flags=re.IGNORECASE,
        )
    )


# --------------------------------------------------------------------------- #
# Line clustering
# --------------------------------------------------------------------------- #

def _cluster_words_into_lines(words: List[dict]) -> List[_Line]:
    """Group word boxes into visual lines using their vertical position.

    ``words`` are pdfplumber word dicts with ``text``, ``x0``, ``x1``, ``top``
    and ``bottom`` keys. Two words belong to the same line when their vertical
    centers are within a tolerance derived from word height.
    """
    if not words:
        return []

    # Sort top-to-bottom, then left-to-right.
    words = sorted(words, key=lambda w: (round(w["top"], 1), w["x0"]))

    lines: List[List[dict]] = []
    current: List[dict] = [words[0]]
    current_top = words[0]["top"]

    for w in words[1:]:
        # Tolerance scales with the word's height so big/small fonts both work.
        height = max(w["bottom"] - w["top"], 1.0)
        tol = height * 0.6
        if abs(w["top"] - current_top) <= tol:
            current.append(w)
        else:
            lines.append(current)
            current = [w]
            current_top = w["top"]
    lines.append(current)

    result: List[_Line] = []
    for group in lines:
        group = sorted(group, key=lambda w: w["x0"])
        text = _clean_text(" ".join(w["text"] for w in group))
        if not text:
            continue
        result.append(
            _Line(
                text=text,
                top=min(w["top"] for w in group),
                bottom=max(w["bottom"] for w in group),
                x0=min(w["x0"] for w in group),
                words=group,
            )
        )
    return result


# --------------------------------------------------------------------------- #
# Header / footer detection
# --------------------------------------------------------------------------- #

def _detect_repeating_headers_footers(pages_lines: List[List[_Line]]) -> set:
    """Return the set of digit-normalized strings that act as headers/footers.

    We look at the top two and bottom two lines of every page; any normalized
    text appearing on most pages is treated as boilerplate (running header,
    footer, or page-number line) and removed.
    """
    n_pages = len(pages_lines)
    if n_pages < 3:
        # Too few pages to reliably tell boilerplate from real content.
        return set()

    from collections import Counter

    counts: Counter = Counter()
    for lines in pages_lines:
        if not lines:
            continue
        candidates = lines[:2] + lines[-2:]
        # Dedupe within a page so a value counts at most once per page.
        seen = set()
        for ln in candidates:
            norm = _normalize_digits(ln.text).strip()
            if norm and norm not in seen:
                seen.add(norm)
                counts[norm] += 1

    threshold = max(2, int(round(n_pages * 0.6)))
    return {norm for norm, c in counts.items() if c >= threshold}


def _strip_boilerplate(lines: List[_Line], boilerplate: set) -> List[_Line]:
    """Drop header/footer/page-number lines from a page's line list."""
    kept: List[_Line] = []
    for ln in lines:
        norm = _normalize_digits(ln.text).strip()
        if norm in boilerplate:
            continue
        if _looks_like_page_number(ln.text):
            continue
        kept.append(ln)
    return kept


# --------------------------------------------------------------------------- #
# Paragraph grouping
# --------------------------------------------------------------------------- #

def _median_line_gap(lines: List[_Line]) -> float:
    """Median vertical gap between consecutive lines on a page."""
    gaps = []
    for prev, cur in zip(lines, lines[1:]):
        gap = cur.top - prev.bottom
        if gap > 0:
            gaps.append(gap)
    if not gaps:
        return 0.0
    return statistics.median(gaps)


def resolve_heading(style: str, pages_lines: List[List["_Line"]]):
    """Resolve a heading-style name to its (regex, max_len) spec.

    For ``"auto"`` we pick the ``legal`` style when at least three distinct
    lines match the legal pattern (a strong sign of a law/regulation), and fall
    back to ``numbered`` otherwise.
    """
    if style in ("numbered", "legal"):
        return HEADING_STYLES[style]
    if style != "auto":
        raise ValueError(f"unknown heading style: {style!r}")

    pat, cap = HEADING_STYLES["legal"]
    distinct = set()
    for lines in pages_lines:
        for ln in lines:
            text = ln.text
            if pat.match(text) and len(text) < cap:
                distinct.add(text)
    return HEADING_STYLES["legal"] if len(distinct) >= 3 else HEADING_STYLES["numbered"]


def is_heading(line, spec) -> bool:
    """True if ``line`` is a heading under ``spec`` (a (regex, max_len) pair)."""
    pat, cap = spec
    text = line.text if isinstance(line, _Line) else line
    return bool(pat.match(text)) and len(text) < cap


def heading_label(text: str):
    """Split a parenthetical legal heading into (clean_label, trailing_body).

    "Article 1 (Purpose) The purpose of ..." -> ("Article 1 (Purpose)",
    "The purpose of ..."). For anything else, returns (text, "").
    """
    m = _HEADING_LABEL_RE.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return text, ""


def _join_lines(lines: List[_Line]) -> str:
    """Join a run of lines into one paragraph, de-hyphenating split words."""
    out = ""
    for ln in lines:
        piece = ln.text
        if not out:
            out = piece
            continue
        m = _HYPHEN_END_RE.search(out)
        if m:
            # Word split across the break: drop the hyphen and fuse the halves.
            out = out[: m.start(1) + 1] + piece.lstrip()
        else:
            out = out + " " + piece
    return _clean_text(out)


def _group_page_paragraphs(lines: List[_Line], gap_factor: float, heading_spec):
    """Group one page's lines into (type, text) paragraph tuples."""
    if not lines:
        return []

    median_gap = _median_line_gap(lines)
    # Threshold above which a vertical gap signals a new paragraph.
    gap_threshold = (median_gap * gap_factor) if median_gap > 0 else 0.0

    paragraphs = []           # list of (kind, text)
    buffer: List[_Line] = []

    def flush():
        if buffer:
            paragraphs.append(("body", _join_lines(buffer)))
            buffer.clear()

    prev: Optional[_Line] = None
    for ln in lines:
        if is_heading(ln, heading_spec):
            flush()
            label, trailing = heading_label(ln.text)
            paragraphs.append(("heading", label))
            if trailing:
                # The sentence trailing a legal heading starts the body para.
                buffer.append(
                    _Line(text=trailing, top=ln.top, bottom=ln.bottom, x0=ln.x0)
                )
            prev = ln
            continue

        if prev is not None:
            gap = ln.top - prev.bottom
            if gap_threshold and gap > gap_threshold:
                flush()
        buffer.append(ln)
        prev = ln

    flush()
    return paragraphs


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #

def extract_paragraphs(
    pdf_path: str, gap_factor: float = 1.6, heading_style: str = "auto"
) -> List[Paragraph]:
    """Extract paragraphs from a text-based PDF.

    Args:
        pdf_path: Path to the PDF file.
        gap_factor: A vertical gap larger than ``median_gap * gap_factor`` is
            treated as a paragraph break. Larger values merge more aggressively.
        heading_style: One of ``"auto"``, ``"numbered"`` or ``"legal"``. ``auto``
            picks ``legal`` for law/regulation-style documents and ``numbered``
            otherwise.

    Returns:
        A list of :class:`Paragraph` in reading order.
    """
    if pdfplumber is None:
        raise RuntimeError(
            "pdfplumber is not installed. Run: pip install -r requirements.txt"
        )

    pages_lines: List[List[_Line]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(
                use_text_flow=False,
                keep_blank_chars=False,
                extra_attrs=["size"],
            )
            pages_lines.append(_cluster_words_into_lines(words))

    boilerplate = _detect_repeating_headers_footers(pages_lines)
    heading_spec = resolve_heading(heading_style, pages_lines)

    paragraphs: List[Paragraph] = []
    current_section = ""
    para_id = 0

    for page_index, raw_lines in enumerate(pages_lines, start=1):
        lines = _strip_boilerplate(raw_lines, boilerplate)
        page_paras = _group_page_paragraphs(lines, gap_factor, heading_spec)

        for kind, text in page_paras:
            text = _clean_text(text)
            if not text:
                continue

            if kind == "heading":
                current_section = text

            # Try to merge with the previous body paragraph across a page break:
            # only when we are at the first paragraph of a new page, the prior
            # paragraph doesn't end a sentence, and this one starts lowercase.
            if (
                kind == "body"
                and paragraphs
                and paragraphs[-1].type == "body"
                and paragraphs[-1].page != page_index
                and page_paras.index((kind, text)) == 0
                and not _SENTENCE_END_RE.search(paragraphs[-1].text)
                and text[:1].islower()
            ):
                prev = paragraphs[-1]
                m = _HYPHEN_END_RE.search(prev.text)
                if m:
                    prev.text = prev.text[: m.start(1) + 1] + text.lstrip()
                else:
                    prev.text = _clean_text(prev.text + " " + text)
                continue

            para_id += 1
            paragraphs.append(
                Paragraph(
                    para_id=para_id,
                    page=page_index,
                    type=kind,
                    section=current_section,
                    text=text,
                )
            )

    return paragraphs


def write_excel(paras: List[Paragraph], out_path: str) -> None:
    """Write extracted paragraphs to a single-sheet Excel workbook.

    Columns: para_id | page | type | section | text. Uses Arial throughout,
    a styled header row, wrapped text for the section/text columns, and a
    frozen header.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "paragraphs"

    headers = ["para_id", "page", "type", "section", "text"]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for p in paras:
        ws.append([p.para_id, p.page, p.type, p.section, p.text])
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.alignment = wrap if col in (4, 5) else top

    # Column widths: fixed for the narrow columns, generous wrap for text.
    widths = {1: 8, 2: 6, 3: 10, 4: 30, 5: 90}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def _default_output(pdf_path: str) -> str:
    import os

    base = os.path.splitext(os.path.basename(pdf_path))[0]
    return f"{base}.xlsx"


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Extract every paragraph from a text-based PDF into Excel."
    )
    parser.add_argument("pdf", help="Path to the input PDF file.")
    parser.add_argument(
        "-o",
        "--output",
        help="Output .xlsx path (default: <pdf-name>.xlsx).",
    )
    parser.add_argument(
        "-g",
        "--gap-factor",
        type=float,
        default=1.6,
        help="Paragraph-break sensitivity: a gap > median*factor splits "
        "paragraphs (default: 1.6).",
    )
    parser.add_argument(
        "--heading-style",
        choices=["auto", "numbered", "legal"],
        default="auto",
        help="Heading detection style (default: auto).",
    )
    args = parser.parse_args(argv)

    out_path = args.output or _default_output(args.pdf)

    try:
        paras = extract_paragraphs(
            args.pdf, gap_factor=args.gap_factor, heading_style=args.heading_style
        )
    except FileNotFoundError:
        print(f"error: file not found: {args.pdf}", file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001 - surface a friendly message
        print(f"error: failed to extract: {exc}", file=sys.stderr)
        return 1

    write_excel(paras, out_path)
    headings = sum(1 for p in paras if p.type == "heading")
    print(
        f"Extracted {len(paras)} paragraphs "
        f"({headings} headings) -> {out_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
