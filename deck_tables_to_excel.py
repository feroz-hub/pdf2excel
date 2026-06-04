"""Extract tables from slide-deck / table-heavy PDFs into Excel.

Where :mod:`pdf_paragraphs_to_excel` reconstructs flowing prose, this module
targets the other common shape: landscape slides and table-heavy pages. For
each page it pulls every *real* table that ``pdfplumber`` can find, stitches
side-by-side boxes (e.g. an "As-Is | To-Be" pair) into one table, and writes
one worksheet per resulting table plus a ``slides_text`` overview sheet.

Public API:
    extract_deck(pdf_path) -> (pages_tables, slides_text)
    write_deck_excel(pages_tables, slides_text, out_path) -> None
    deck_to_excel(pdf_path, out_path) -> None     # convenience: extract + write

CLI:
    python deck_tables_to_excel.py deck.pdf -o out.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from typing import List, Tuple

try:
    import pdfplumber
except ImportError:  # pragma: no cover - dependency hint
    pdfplumber = None

_CID_RE = re.compile(r"\(cid:\d+\)")
_NAVY = "2F5496"
_MAX_COL_WIDTH = 60

# Horizontal slack (pts) allowed when deciding two tables sit side by side.
_X_SLACK = 2.0
# Minimum fraction of the shorter table's height that must overlap vertically.
_V_OVERLAP_MIN = 0.5


@dataclass
class _Table:
    """A cleaned table: rows of strings plus its bounding box on the page."""

    rows: List[List[str]]
    bbox: Tuple[float, float, float, float]  # (x0, top, x1, bottom)

    @property
    def x0(self) -> float:
        return self.bbox[0]

    @property
    def top(self) -> float:
        return self.bbox[1]

    @property
    def x1(self) -> float:
        return self.bbox[2]

    @property
    def bottom(self) -> float:
        return self.bbox[3]

    @property
    def height(self) -> float:
        return self.bottom - self.top

    @property
    def width(self) -> int:
        """Number of columns (widest row)."""
        return max((len(r) for r in self.rows), default=0)


# --------------------------------------------------------------------------- #
# Cell / table cleaning
# --------------------------------------------------------------------------- #

def _clean_cell(value) -> str:
    """Normalize a raw table cell: drop cid artifacts, collapse whitespace."""
    if value is None:
        return ""
    text = _CID_RE.sub("", str(value))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _is_real_table(rows: List[List[str]]) -> bool:
    """Keep tables with >=2 rows and >=2 non-empty cells.

    pdfplumber reports full-page border rectangles as single-cell "tables";
    this filter drops those while keeping genuine grids.
    """
    if len(rows) < 2:
        return False
    non_empty = sum(1 for row in rows for cell in row if cell)
    return non_empty >= 2


# --------------------------------------------------------------------------- #
# Side-by-side merging
# --------------------------------------------------------------------------- #

def _vertical_overlap_fraction(a: _Table, b: _Table) -> float:
    """Fraction of the shorter table's height that overlaps vertically."""
    overlap = min(a.bottom, b.bottom) - max(a.top, b.top)
    if overlap <= 0:
        return 0.0
    shorter = min(a.height, b.height)
    if shorter <= 0:
        return 0.0
    return overlap / shorter


def _concat_columnwise(group: List[_Table]) -> _Table:
    """Concatenate a left-to-right group of tables column-wise into one table."""
    n_rows = max(len(t.rows) for t in group)
    widths = [t.width for t in group]
    merged: List[List[str]] = []
    for k in range(n_rows):
        row: List[str] = []
        for t, w in zip(group, widths):
            cells = list(t.rows[k]) if k < len(t.rows) else []
            cells += [""] * (w - len(cells))   # pad short rows
            row.extend(cells)
        merged.append(row)

    bbox = (
        min(t.x0 for t in group),
        min(t.top for t in group),
        max(t.x1 for t in group),
        max(t.bottom for t in group),
    )
    return _Table(rows=merged, bbox=bbox)


def _merge_side_by_side(tables: List[_Table]) -> List[_Table]:
    """Stitch horizontally adjacent, vertically aligned tables into one each."""
    if len(tables) < 2:
        return tables

    ordered = sorted(tables, key=lambda t: t.x0)
    groups: List[List[_Table]] = []
    current: List[_Table] = [ordered[0]]

    for nxt in ordered[1:]:
        cur = current[-1]
        side_by_side = nxt.x0 >= cur.x1 - _X_SLACK
        aligned = _vertical_overlap_fraction(cur, nxt) > _V_OVERLAP_MIN
        if side_by_side and aligned:
            current.append(nxt)
        else:
            groups.append(current)
            current = [nxt]
    groups.append(current)

    return [g[0] if len(g) == 1 else _concat_columnwise(g) for g in groups]


# --------------------------------------------------------------------------- #
# Extraction
# --------------------------------------------------------------------------- #

def _extract_page_tables(page) -> List[_Table]:
    """Return the cleaned, real tables found on a single page."""
    result: List[_Table] = []
    for tbl in page.find_tables():
        raw = tbl.extract()
        rows = [[_clean_cell(c) for c in row] for row in raw]
        if not _is_real_table(rows):
            continue
        result.append(_Table(rows=rows, bbox=tuple(tbl.bbox)))
    return result


def extract_deck(pdf_path: str):
    """Extract tables and slide text from a deck/table PDF.

    Returns:
        (pages_tables, slides_text) where ``pages_tables`` is a list of
        ``(page_no, [Table, ...])`` and ``slides_text`` is a list of
        ``(page_no, text)``.
    """
    if pdfplumber is None:
        raise RuntimeError(
            "pdfplumber is not installed. Run: pip install -r requirements.txt"
        )

    pages_tables = []
    slides_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            tables = _merge_side_by_side(_extract_page_tables(page))
            pages_tables.append((page_no, tables))
            text = _CID_RE.sub("", page.extract_text() or "")
            slides_text.append((page_no, text.strip()))
    return pages_tables, slides_text


# --------------------------------------------------------------------------- #
# Excel output
# --------------------------------------------------------------------------- #

def _style_sheet(ws, n_cols: int) -> None:
    """Apply Arial, navy header, frozen header, wrap and autosized columns."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_NAVY)
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap

    # Autosize columns based on the longest line in each cell, capped.
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        longest = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value is None:
                    continue
                for line in str(cell.value).splitlines() or [""]:
                    longest = max(longest, len(line))
        ws.column_dimensions[letter].width = min(longest + 2, _MAX_COL_WIDTH)

    ws.freeze_panes = "A2"


def _unique_title(wb, base: str) -> str:
    """Return a worksheet title based on ``base`` that is unique and <=31 chars."""
    base = base[:31]
    if base not in wb.sheetnames:
        return base
    n = 2
    while True:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        if title not in wb.sheetnames:
            return title
        n += 1


def write_deck_excel(pages_tables, slides_text, out_path: str) -> None:
    """Write extracted deck tables (one sheet each) plus a slides_text sheet."""
    from openpyxl import Workbook

    wb = Workbook()
    # Drop the default sheet; we add our own and rename appropriately.
    default = wb.active
    wb.remove(default)

    for page_no, tables in pages_tables:
        for idx, table in enumerate(tables):
            base = f"slide{page_no}" if idx == 0 else f"slide{page_no}_{idx + 1}"
            ws = wb.create_sheet(title=_unique_title(wb, base))
            width = table.width
            for row in table.rows:
                padded = list(row) + [""] * (width - len(row))
                ws.append(padded)
            _style_sheet(ws, width)

    # slides_text overview: one row per slide. When there were no tables this is
    # the only (and therefore first/visible) sheet.
    ws_text = wb.create_sheet(title="slides_text")
    ws_text.append(["slide", "text"])
    for page_no, text in slides_text:
        ws_text.append([page_no, text])
    _style_text_sheet(ws_text)

    wb.save(out_path)


def _style_text_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_NAVY)
    body_font = Font(name="Arial", size=10)

    for col in (1, 2):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for row in ws.iter_rows(min_row=2):
        row[0].font = body_font
        row[0].alignment = Alignment(vertical="top", horizontal="center")
        row[1].font = body_font
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = _MAX_COL_WIDTH
    ws.freeze_panes = "A2"


def deck_to_excel(pdf_path: str, out_path: str) -> Tuple[int, int]:
    """Extract a deck PDF and write it to Excel. Returns (n_tables, n_pages)."""
    pages_tables, slides_text = extract_deck(pdf_path)
    write_deck_excel(pages_tables, slides_text, out_path)
    n_tables = sum(len(tables) for _, tables in pages_tables)
    return n_tables, len(slides_text)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Extract tables from a slide-deck / table-heavy PDF into Excel."
    )
    parser.add_argument("pdf", help="Path to the input PDF file.")
    parser.add_argument("-o", "--output", help="Output .xlsx path.")
    args = parser.parse_args(argv)

    out_path = args.output or (
        os.path.splitext(os.path.basename(args.pdf))[0] + ".xlsx"
    )

    try:
        n_tables, n_pages = deck_to_excel(args.pdf, out_path)
    except FileNotFoundError:
        print(f"error: file not found: {args.pdf}", file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001
        print(f"error: failed to extract: {exc}", file=sys.stderr)
        return 1

    print(f"Extracted {n_tables} tables from {n_pages} pages -> {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
